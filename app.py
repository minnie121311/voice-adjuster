from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import os
import csv
import json
from datetime import datetime
import random
import parselmouth
from parselmouth.praat import call
import io
import threading
import base64
import requests
try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export disabled.")


app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-change-this-in-production')


# 연구자 설정
RESEARCHER_EMAIL = 'minnie1211@gmail.com'
SMTP_EMAIL = 'minnie1211@gmail.com'
SMTP_PASSWORD = 'apck becz medp lddg'

# Prolific 완료 코드 - Prolific에서 study를 만들면 발급되는 completion code.
# 값이 설정되기 전까지는 thankyou 페이지에 "Return to Prolific" 버튼이 표시되지 않음.
PROLIFIC_COMPLETION_CODE = os.environ.get('PROLIFIC_COMPLETION_CODE', '')


# 영구 저장 위치 (Railway Volume을 이 경로에 마운트하고 DATA_ROOT 환경변수로 지정하면
# 재배포되어도 참가자 데이터가 유지됨. 미설정 시 기존처럼 앱 디렉터리에 저장되어
# 재배포할 때마다 초기화됨 - Railway는 컨테이너 파일시스템을 임시로 취급함)
DATA_ROOT = os.environ.get('DATA_ROOT', '.')

# CSV 파일들
ALL_DATA_CSV = os.path.join(DATA_ROOT, 'all_participant_data.csv')
PHASE2_CSV = os.path.join(DATA_ROOT, 'phase2_data.csv')
PHASE2_FOLDERS = ['p1', 'p2', 'p3', 'n1', 'n2', 'n3', 'i1', 'i2', 'i3']
OUTPUT_DIR = 'static/adjusted'
DATA_DIR = os.path.join(DATA_ROOT, 'data', 'phase1')


# 폴더 생성
for folder in [DATA_ROOT, OUTPUT_DIR, DATA_DIR]:
    if not os.path.exists(folder):
        os.makedirs(folder)


# 통합 CSV 초기화
ALL_DATA_HEADER = [
    'session_id', 'timestamp', 'data_type',
    'consent_given', 'consent_time',
    'lsas_fear', 'lsas_avoidance', 'lsas_total',
    'phase1_audio', 'phase1_trustworthiness', 'phase1_anxiety',
    'phase1_preference', 'phase1_dominance', 'phase1_warmth', 'phase1_listen_time',
    'phase2_folder', 'phase2_formant', 'phase2_pitch',
    'prolific_id', 'age'
]

if not os.path.exists(ALL_DATA_CSV):
    with open(ALL_DATA_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(ALL_DATA_HEADER)

if not os.path.exists(PHASE2_CSV):
    with open(PHASE2_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['timestamp', 'session_id', 'folder_name', 'file_index', 'formant_value', 'pitch_semitones'])



# ===== 실험 흐름 라우트 =====

@app.route('/')
def home():
    return redirect(url_for('welcome'))


@app.route('/welcome')
def welcome():
    # Prolific redirects here with ?PROLIFIC_PID=...&STUDY_ID=...&SESSION_ID=...
    # Capture it once into the session so later pages don't need the query string.
    prolific_pid = request.args.get('PROLIFIC_PID')
    if prolific_pid:
        session['prolific_pid'] = prolific_pid
    return render_template('welcome.html')


@app.route('/consent')
def consent():
    return render_template('consent.html')


@app.route('/submit-consent', methods=['POST'])
def submit_consent():
    try:
        session['consent_given'] = True
        session['consent_time'] = datetime.now().isoformat()

        if 'study_session_id' not in session:
            session['study_session_id'] = datetime.now().strftime('%Y%m%d_%H%M%S_%f')

        with open(ALL_DATA_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                session['study_session_id'],
                datetime.now().isoformat(),
                'consent',
                True,
                session['consent_time'],
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''
            ])

        print(f"Consent saved: {session['study_session_id']}")

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/personal-info')
def personal_info():
    return render_template('personal_info.html', prolific_pid=session.get('prolific_pid', ''))

@app.route('/submit-personal-info', methods=['POST'])
def submit_personal_info():
    try:
        data = request.json or {}
        prolific_id = (data.get('prolific_id') or '').strip()
        age = (data.get('age') or '').strip()

        if not prolific_id or not age:
            return jsonify({'error': 'prolific_id and age are required'}), 400

        session['prolific_id'] = prolific_id
        session['age'] = age

        if 'study_session_id' not in session:
            session['study_session_id'] = datetime.now().strftime('%Y%m%d_%H%M%S_%f')

        with open(ALL_DATA_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                session['study_session_id'],
                datetime.now().isoformat(),
                'personal_info',
                '', '',
                '', '', '',
                '', '', '', '', '', '', '',
                '', '', '',
                prolific_id, age
            ])

        print(f"Personal info saved: {session['study_session_id']} prolific_id={prolific_id}")

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/lsas')
def lsas():
    return render_template('lsas.html')


@app.route('/submit-lsas', methods=['POST'])
def submit_lsas():
    try:
        data = request.json
        responses = data.get('responses', [])

        fear_total = sum(r['fear'] for r in responses)
        avoidance_total = sum(r['avoidance'] for r in responses)
        total_score = fear_total + avoidance_total

        session['lsas_fear'] = fear_total
        session['lsas_avoidance'] = avoidance_total
        session['lsas_total'] = total_score

        with open(ALL_DATA_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                session.get('study_session_id'),
                datetime.now().isoformat(),
                'lsas',
                '', '',
                fear_total, avoidance_total, total_score,
                '', '', '', '', '', '', '', '', '', '',
                '', ''
            ])

        print(f"LSAS saved: Fear={fear_total}, Avoidance={avoidance_total}, Total={total_score}")

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error in submit_lsas: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/phase1')
def phase1():
    if 'participant_id' not in session:
        session['participant_id'] = f"P{datetime.now().strftime('%Y%m%d%H%M%S')}"
    return render_template('phase1.html')


@app.route('/api/submit_phase1', methods=['POST'])
def submit_phase1():
    try:
        data = request.json
        participant_id = data.get('participant_id', 'anonymous')
        responses = data.get('responses', [])

        with open(ALL_DATA_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for resp in responses:
                writer.writerow([
                    session.get('study_session_id', participant_id),
                    resp.get('timestamp'),
                    'phase1',
                    '', '',
                    '', '', '',
                    resp.get('filename'),
                    resp.get('trustworthiness'),
                    resp.get('anxiety'),
                    resp.get('preference'),
                    resp.get('dominance'),
                    resp.get('warmth'),
                    resp.get('listenTime', 0),
                    '', '', '',
                    '', ''
                ])

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{participant_id}_{timestamp}.json"
        filepath = os.path.join(DATA_DIR, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"Phase 1 data saved: {filename}")

        threading.Thread(target=send_phase1_email, args=(participant_id, len(responses)), daemon=True).start()

        return jsonify({
            'success': True,
            'message': 'Data saved successfully',
            'participant_id': participant_id
        })

    except Exception as e:
        print(f"Error in submit_phase1: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


def send_email_resend(to, subject, body, attachment_bytes=None, attachment_name=None):
    try:
        api_key = os.environ.get('RESEND_API_KEY', '').strip()
        if not api_key:
            print('RESEND_API_KEY not set')
            return False, 'RESEND_API_KEY not set'
        payload = {
            'from': 'Voice Study <onboarding@resend.dev>',
            'to': [to],
            'subject': subject,
            'text': body
        }
        if attachment_bytes and attachment_name:
            payload['attachments'] = [{
                'filename': attachment_name,
                'content': base64.b64encode(attachment_bytes).decode()
            }]
        resp = requests.post(
            'https://api.resend.com/emails',
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            json=payload,
            timeout=15
        )
        print(f'Resend response: {resp.status_code} {resp.text}')
        return resp.status_code == 200, f'{resp.status_code}: {resp.text}'
    except Exception as e:
        print(f'Resend exception: {e}')
        return False, str(e)


def send_phase1_email(participant_id, response_count):
    try:
        print(f"Sending Phase 1 email for: {participant_id}")
        body = f'''A participant has completed Phase 1 (Voice Evaluation).

Participant ID: {participant_id}
Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Responses: {response_count} audio evaluations

Data will be included in final Excel file after Phase 2 completion.'''
        ok, _ = send_email_resend(RESEARCHER_EMAIL, f'Voice Study - Phase 1 Complete - {participant_id}', body)
    except Exception as e:
        print(f"Phase 1 email error: {str(e)}")


@app.route('/phase2')
def phase2():
    return render_template('phase2.html')


@app.route('/start-phase2', methods=['POST'])
def start_phase2():
    try:
        shuffled = PHASE2_FOLDERS.copy()
        random.shuffle(shuffled)

        session['phase2_folders'] = shuffled
        session['current_index'] = 0

        if 'session_id' not in session:
            session['session_id'] = datetime.now().strftime('%Y%m%d_%H%M%S_%f')

        print(f"Phase 2 started: {session['session_id']}")

        return jsonify({
            'current_folder': shuffled[0],
            'index': 0,
            'total': len(shuffled)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/adjust-voice', methods=['POST'])
def adjust_voice():
    try:
        data = request.json
        folder = data.get('folder')
        formant = int(data.get('formant', 0))
        pitch_semitones = float(data.get('pitch', 0))  # 🔥 int → float

        # 🔥 범위 제한 -5 ~ +5
        pitch_semitones = max(-5, min(5, pitch_semitones))

        formant_file = f'static/audio/{folder}/{folder}_{formant}.wav'

        if not os.path.exists(formant_file):
            return jsonify({'error': f'File not found'}), 404

        # 🔥 0.01 미만은 원본 반환
        if abs(pitch_semitones) < 0.01:
            return jsonify({
                'output_file': f'audio/{folder}/{folder}_{formant}.wav',
                'parameters': {'formant': formant, 'pitch': 0}
            })

        sound = parselmouth.Sound(formant_file)
        pitch_factor = 2.0 ** (pitch_semitones / 12.0)

        try:
            manipulation = call(sound, "To Manipulation", 0.01, 75, 600)
            pitch_tier = call(manipulation, "Extract pitch tier")
            call(pitch_tier, "Multiply frequencies", sound.xmin, sound.xmax, pitch_factor)
            call([pitch_tier, manipulation], "Replace pitch tier")
            sound = call(manipulation, "Get resynthesis (overlap-add)")
        except:
            return jsonify({
                'output_file': f'audio/{folder}/{folder}_{formant}.wav',
                'parameters': {'formant': formant, 'pitch': 0}
            })

        max_val = max(abs(sound.get_value(t)) for t in sound.xs())
        if max_val > 0.95:
            sound = sound * (0.95 / max_val)

        # 🔥 파일명에 소수점 포함
        output_filename = f"{folder}_f{formant}_p{pitch_semitones:.1f}.wav"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        sound.save(output_path, 'WAV')

        return jsonify({
            'output_file': f'adjusted/{output_filename}',
            'parameters': {'formant': formant, 'pitch': pitch_semitones}
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/play/<path:filepath>')
def play_audio(filepath):
    try:
        file_path = f'static/{filepath}'
        if os.path.exists(file_path):
            return send_file(file_path, mimetype='audio/wav')
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/submit-adjustment', methods=['POST'])
def submit_adjustment():
    try:
        data = request.json
        folders = session.get('phase2_folders', [])
        index = session.get('current_index', 0)

        with open(PHASE2_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                datetime.now().isoformat(),
                session.get('session_id'),
                data.get('folder'),
                index,
                data.get('formant'),
                data.get('pitch')
            ])

        with open(ALL_DATA_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                session.get('study_session_id', session.get('session_id')),
                datetime.now().isoformat(),
                'phase2',
                '', '',
                '', '', '',
                '', '', '', '', '', '', '',
                data.get('folder'),
                data.get('formant'),
                data.get('pitch'),
                '', ''
            ])

        session['current_index'] = index + 1

        if session['current_index'] >= len(folders):
            sid = session.get('study_session_id')
            threading.Thread(target=send_complete_excel, args=(sid,), daemon=True).start()

            return jsonify({
                'completed': True,
                'message': 'Study completed! Thank you.'
            })

        return jsonify({
            'success': True,
            'next_folder': folders[session['current_index']],
            'index': session['current_index'],
            'total': len(folders)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def build_master_workbook():
    """Rebuild a wide-format workbook covering every participant found in ALL_DATA_CSV."""
    from collections import OrderedDict

    all_rows = []
    if os.path.exists(ALL_DATA_CSV):
        with open(ALL_DATA_CSV, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            all_rows = list(reader)

    sessions = OrderedDict()
    for row in all_rows:
        sessions.setdefault(row[0], []).append(row)

    wb = Workbook()

    # --- Sheet 1: Payment (Prolific ID + age, for payment reconciliation) ---
    ws0 = wb.active
    ws0.title = 'Payment'
    ws0.append(['Session ID', 'Prolific ID', 'Age'])
    for sid, rows in sessions.items():
        pi_row = next((r for r in rows if r[2] == 'personal_info'), None)
        prolific_id = pi_row[18] if pi_row else ''
        age = pi_row[19] if pi_row else ''
        ws0.append([sid, prolific_id, age])

    # --- Sheet 2: Phase 1 wide (all participants) ---
    ws1 = wb.create_sheet('Phase 1')
    all_p1_files = []
    seen = set()
    for rows in sessions.values():
        for r in sorted([x for x in rows if x[2] == 'phase1'], key=lambda x: x[8]):
            fname = r[8].replace('.mp3', '')
            if fname not in seen:
                seen.add(fname)
                all_p1_files.append(fname)
    p1_header = ['Session ID']
    for fname in all_p1_files:
        for m in ['trust', 'anxiety', 'pref', 'dom', 'warmth']:
            p1_header.append(f'{fname}_{m}')
    ws1.append(p1_header)
    for sid, rows in sessions.items():
        p1_map = {r[8].replace('.mp3', ''): r for r in rows if r[2] == 'phase1'}
        row_data = [sid]
        for fname in all_p1_files:
            if fname in p1_map:
                r = p1_map[fname]
                row_data.extend([r[9], r[10], r[11], r[12], r[13]])
            else:
                row_data.extend(['', '', '', '', ''])
        ws1.append(row_data)

    # --- Sheet 2: Phase 2 wide (all participants) ---
    ws2 = wb.create_sheet('Phase 2')
    all_p2_folders = []
    seen2 = set()
    for rows in sessions.values():
        for r in sorted([x for x in rows if x[2] == 'phase2'], key=lambda x: x[15]):
            folder = r[15]
            if folder not in seen2:
                seen2.add(folder)
                all_p2_folders.append(folder)
    p2_header = ['Session ID']
    for folder in all_p2_folders:
        p2_header.append(f'{folder}_formant')
        p2_header.append(f'{folder}_pitch')
    ws2.append(p2_header)
    for sid, rows in sessions.items():
        p2_map = {r[15]: r for r in rows if r[2] == 'phase2'}
        row_data = [sid]
        for folder in all_p2_folders:
            if folder in p2_map:
                r = p2_map[folder]
                row_data.extend([r[16], r[17]])
            else:
                row_data.extend(['', ''])
        ws2.append(row_data)

    # --- Sheet 3: LSAS ---
    ws3 = wb.create_sheet('LSAS')
    ws3.append(['Session ID', 'Timestamp', 'Fear Score', 'Avoidance Score', 'Total Score'])
    for rows in sessions.values():
        for r in rows:
            if r[2] == 'lsas':
                ws3.append([r[0], r[1], r[5], r[6], r[7]])

    # --- Sheet 4: Consent ---
    ws4 = wb.create_sheet('Consent')
    ws4.append(['Session ID', 'Timestamp', 'Consent Given', 'Consent Time'])
    for rows in sessions.values():
        for r in rows:
            if r[2] == 'consent':
                ws4.append([r[0], r[1], r[3], r[4]])

    return wb, len(sessions)


def send_complete_excel(session_id):
    try:
        if not session_id:
            print("send_complete_excel: session_id is None, skipping")
            return

        print(f"Rebuilding master Excel after completion: {session_id}")

        if not EXCEL_AVAILABLE:
            print("Excel export not available")
            return

        wb, total_sessions = build_master_workbook()

        master_filename = 'all_participants_master.xlsx'
        master_path = os.path.join(DATA_DIR, master_filename)
        wb.save(master_path)

        body = f'''A participant has completed the entire study!

Just completed - Session ID: {session_id}
Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Total participants so far: {total_sessions}

Attached is the updated master Excel file covering ALL participants to date, with tabs:
- Phase 1 / Phase 2 / LSAS / Consent'''

        with open(master_path, 'rb') as f:
            excel_bytes = f.read()

        ok, detail = send_email_resend(
            RESEARCHER_EMAIL,
            f'Voice Study - Master Data Updated ({total_sessions} participants)',
            body,
            attachment_bytes=excel_bytes,
            attachment_name=master_filename
        )
        if ok:
            print(f"✓ Master Excel email sent to {RESEARCHER_EMAIL}")
        else:
            print(f"Master Excel email failed via Resend: {detail}")

    except Exception as e:
        print(f"Excel email error: {str(e)}")
        import traceback
        traceback.print_exc()



@app.route('/admin/test-email')
def test_email():
    admin_key = request.args.get('key')
    if admin_key != 'ucl-voice-study-2026':
        return jsonify({'error': 'Unauthorized'}), 403

    ok, detail = send_email_resend(
        RESEARCHER_EMAIL,
        'Voice Study - Email Test',
        f'Test email from Railway at {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    )
    return jsonify({'status': 'success' if ok else 'failed', 'detail': detail})


@app.route('/admin/csv-status')
def csv_status():
    admin_key = request.args.get('key')
    if admin_key != 'ucl-voice-study-2026':
        return jsonify({'error': 'Unauthorized'}), 403

    rows = 0
    last_session = None
    if os.path.exists(ALL_DATA_CSV):
        with open(ALL_DATA_CSV, 'r') as f:
            lines = f.readlines()
        rows = len(lines) - 1
        if rows > 0:
            last_session = lines[-1].split(',')[0]

    excel_files = []
    if os.path.exists(DATA_DIR):
        excel_files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx')]

    return jsonify({
        'csv_rows': rows,
        'last_session': last_session,
        'excel_files': excel_files
    })


@app.route('/admin/reset-data', methods=['POST'])
def reset_data():
    admin_key = request.args.get('key')
    if admin_key != 'ucl-voice-study-2026':
        return jsonify({'error': 'Unauthorized'}), 403

    with open(ALL_DATA_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(ALL_DATA_HEADER)

    with open(PHASE2_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['timestamp', 'session_id', 'folder_name', 'file_index', 'formant_value', 'pitch_semitones'])

    removed = 0
    if os.path.exists(DATA_DIR):
        for fname in os.listdir(DATA_DIR):
            fpath = os.path.join(DATA_DIR, fname)
            if os.path.isfile(fpath):
                os.remove(fpath)
                removed += 1

    print(f"Data reset via admin endpoint, {removed} files removed from {DATA_DIR}")
    return jsonify({'success': True, 'files_removed': removed})


@app.route('/thankyou')
def thankyou():
    prolific_url = None
    if PROLIFIC_COMPLETION_CODE:
        prolific_url = f'https://app.prolific.com/submissions/complete?cc={PROLIFIC_COMPLETION_CODE}'
    return render_template('thankyou.html', prolific_url=prolific_url)


@app.route('/download-csv')
def download_csv():
    admin_key = request.args.get('key')
    if admin_key != 'ucl-voice-study-2026':
        return jsonify({'error': 'Unauthorized'}), 403

    try:
        if os.path.exists(ALL_DATA_CSV):
            return send_file(ALL_DATA_CSV, as_attachment=True, download_name='all_study_data.csv')
        else:
            return jsonify({'error': 'No data available'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500




@app.route('/admin/download-excel')
def download_excel():
    admin_key = request.args.get('key')
    if admin_key != 'ucl-voice-study-2026':
        return jsonify({'error': 'Unauthorized'}), 403

    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500

    try:
        wb, total_sessions = build_master_workbook()

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'voice_study_all_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/report-issue', methods=['POST'])
def report_issue():
    data = request.get_json()
    message = data.get('message', '')
    page = data.get('page', 'unknown')
    body = f'''Issue reported from page: {page}

Message:
{message}

Time: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'''
    ok, _ = send_email_resend(RESEARCHER_EMAIL, f'Voice Study - Issue Report ({page})', body)
    return jsonify({'ok': ok})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    print(f"Starting Flask app on port {port}...")
    print(f"Excel export: {'Enabled' if EXCEL_AVAILABLE else 'Disabled (install openpyxl)'}")
    app.run(host='0.0.0.0', port=port, debug=True)
